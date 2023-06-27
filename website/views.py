from flask import Blueprint, render_template,request
import matplotlib
matplotlib.use('agg')



views = Blueprint("views", __name__)

@views.route("/")

def home():
    return render_template("base.html")


@views.route("/map")
def map():
    import geopandas as gpd
    import json


    folder_path = r'D:\Trabalho de Licenciatura\New folder\Python\Moz Geo Data\Distritos'


    shapefile_path = folder_path + "\Distritos.shp"


    districts = gpd.read_file(shapefile_path)


    districts_json = districts.to_crs(epsg='4326').to_json()

    return render_template("map.html", districts_json=districts_json)



@views.route("/map/district/<district_name>")
def district_page(district_name):
    import pulp 
    import openpyxl
    from flask import jsonify

    
    book= openpyxl.load_workbook(r"D:\Trabalho de Licenciatura\New folder\Python\WebApp\website\static\Dados_py.xlsx")
    worksheet=book["Dados"]
    
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if row[0]==district_name:
            A=str(row[1])
            B=str(row[2])
            C=str(row[3])

    A=float(A)
    B=float(B)
    C=float(C)
    
    return render_template("district.html", district_name=district_name,A=A, B=B, C=C)


@views.route("/map/district/<district_name>/calculate", methods=["POST"])
def calculate(district_name):
    import pulp 
    import openpyxl
    from flask import jsonify
    import matplotlib.pyplot as plt
    import numpy as np
    import mpl_toolkits.mplot3d as Axes3D
    
    demand = float(request.form["demand"])
    
    book = openpyxl.load_workbook(r"D:\Trabalho de Licenciatura\New folder\Python\WebApp\website\static\Dados_py.xlsx")
    worksheet = book["Dados"]
    
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if row[0] == district_name:
            A = str(row[1])
            B = str(row[2])
            C = str(row[3])
    
    model = pulp.LpProblem(name="Lp", sense=pulp.const.LpMinimize)
    A = float(A)
    B = float(B)
    C = float(C)
    x_1 = pulp.LpVariable(name="Solar", lowBound=0)
    x_2 = pulp.LpVariable(name="Eólica", lowBound=0)
    x_3 = pulp.LpVariable(name="Hídrica", lowBound=0)
    
    obj_func = A * x_1 + B * x_2 + C * x_3
    
    model += obj_func
    
    model += (x_1 + x_2 + x_3 >= demand)
    model += (41*x_1+11*x_2+24*x_3 <= 20*demand)  
    model += ( 0.309*x_2 >= x_3)
    model += (0.399*x_1 >= x_3) 
    model += (0.744*x_2 >= x_1)
    
    
    results_dict = {}
    status = model.solve()
    for var in model.variables():
        results_dict[var.name] = var.value()
    
    x_1_value = results_dict["Solar"]
    x_2_value = results_dict["Eólica"]
    x_3_value = results_dict["Hídrica"]
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')

    x = np.linspace(0, demand, 1000)
    y = np.linspace(0, demand, 1000)
    X, Y = np.meshgrid(x, y)

    Z1 = demand - X - Y
    Z2 = (20*demand - 41*X - 11 * Y) / 24
    Z3 = (0.399*X)
    Z4 = (0.309 * Y)

    ax.scatter(x_1_value, x_2_value, x_3_value, color='red', label='Optimal Solution')    
    ax.plot_surface(X, Y, Z1, alpha=0.5, color='skyblue', label=f'$x_1$ + $x_2$ + $x_3$ = {demand}')
    ax.plot_surface(X, Y, Z2, alpha=0.5, color='green', label=f'41$x_1$+11$x_2$+24$x_3$ = 20*{demand}')
    ax.plot_surface(X, Y, Z3, alpha=0.5, color='yellow', label=f'$x_3$=0.399$x_1$ ')
    ax.plot_surface(X, Y, Z4, alpha=0.5, color='cyan', label=f'$x_3$=0.309$x_2$ ')  
    
    ax.set_xlabel('$x_{1}$ (Solar)')
    ax.set_ylabel('$x_{2}$ (Eólica)')
    ax.set_zlabel('$x_{3}$ (Hídrica)')


    legend_handles = [plt.Rectangle((0, 0), 1, 1, fc='skyblue', alpha=0.5),
                      plt.Rectangle((0, 0), 0.5, 0.5, fc='green', alpha=0.5),plt.Rectangle((0, 0), 1, 1, fc='yellow', alpha=0.5),
                      plt.Rectangle((0, 0), 1, 1, fc='cyan', alpha=0.5),
                      plt.Line2D([0], [0], marker='o', color='red', label='Solução óptima')]
    legend_labels = [f'$x_1$ + $x_2$ + $x_3$ = {demand}',
                     f'41$x_1$+11$x_2$+24$x_3$ = 20*{demand}',f'$x_3$=0.399$x_1$ ',f'$x_3$=0.309$x_2$ ',
                     'Solução óptima']


    ax.legend(handles=legend_handles, labels=legend_labels, loc='upper left',
              bbox_to_anchor=(0.05, 0.99), ncol=2, prop={'size': 7})
    

    plot_image = 'D:\\Trabalho de Licenciatura\\New folder\\Python\\WebApp\\website\\static\\images\\plot.png'
    plt.savefig(plot_image)


    plt.close()


    return jsonify(results_dict=results_dict, plot_image=plot_image)
